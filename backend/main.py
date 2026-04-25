import os
import shutil
from datetime import datetime, timezone, timedelta
from typing import Optional
try:
    from dotenv import load_dotenv
    load_dotenv(os.path.join(os.path.dirname(__file__), ".env"))
except ImportError:
    pass  # On Render, env vars are injected — dotenv not needed
from fastapi import Depends, FastAPI, File, HTTPException, UploadFile, status, Body
from fastapi.security import OAuth2PasswordBearer, HTTPBearer
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from fastapi.openapi.utils import get_openapi
from pydantic import BaseModel
from sqlalchemy.orm import Session

try:
    from .auth import ALGORITHM, SECRET_KEY, clear_otp, create_access_token, hash_password, send_otp_email, send_otp_email_change, verify_password, verify_stored_otp
    from .database import get_session, init_db
    from . import models
except ImportError:
    from auth import ALGORITHM, SECRET_KEY, clear_otp, create_access_token, hash_password, send_otp_email, send_otp_email_change, verify_password, verify_stored_otp
    from database import get_session, init_db
    import models
from jose import JWTError, jwt


init_db()

tags_metadata = [
    {"name": "auth",         "description": "Authentication — login, register, OTP, password reset, FCM token."},
    {"name": "profile",      "description": "User profiles — view, update, email change, resume upload."},
    {"name": "admin",        "description": "Admin operations — list users, bulk register, delete accounts."},
    {"name": "alumni",       "description": "Alumni directory — search and filter alumni profiles."},
    {"name": "internships",  "description": "Internship board — post, browse, apply, manage applications."},
    {"name": "applications", "description": "Internship applications — track status, shortlist, select."},
    {"name": "messages",     "description": "Direct messaging — send, edit, delete, bulk delete, conversations."},
    {"name": "events",       "description": "Events & announcements — post, list, delete."},
    {"name": "connections",  "description": "Connection requests — send, accept, reject between students and alumni."},
    {"name": "resources",    "description": "Educational resources — upload documents/links, list, delete."},
    {"name": "common",       "description": "File upload utility endpoint."},
]

app = FastAPI(
    title="Alumni Network API",
    description="""
## Alumni-Student Network Platform — REST API

Built with **FastAPI** + **SQLAlchemy**. Supports JWT authentication, FCM push notifications, file uploads via Cloudinary, and real-time messaging.

### Authentication
All protected endpoints require a **Bearer JWT token**.

To get a token:
1. `POST /auth/login` with `email`, `password`, `role`
2. Copy the `access_token` from the response
3. Click **Authorize** (🔒) above and enter: `Bearer <your_token>`

### Roles
| Role | Permissions |
|------|-------------|
| `admin` | Full access — manage users, events, resources, notifications |
| `staff` | Post events/resources, message all users, send notifications |
| `alumni` | Post internships, connect with students, message connected users |
| `student` | Apply to internships, connect with alumni, message connected alumni + admin/staff |
""",
    version="2.0.0",
    openapi_tags=tags_metadata,
    contact={
        "name": "Alumni Network Support",
        "email": "gajagopi2006@gmail.com",
    },
    license_info={
        "name": "Private",
    },
    swagger_ui_parameters={
        "defaultModelsExpandDepth": -1,   # hide schemas section by default
        "docExpansion": "none",            # collapse all endpoints by default
        "filter": True,                    # enable search/filter bar
        "persistAuthorization": True,      # keep auth token across page refreshes
    },
)
IS_VERCEL = os.environ.get("VERCEL") == "1"

if not IS_VERCEL:
    UPLOAD_DIR = os.path.join(os.path.dirname(__file__), "uploads")
    
    # Only create locally
    os.makedirs(UPLOAD_DIR, exist_ok=True)

    app.mount("/uploads", StaticFiles(directory=UPLOAD_DIR), name="uploads")
else:
    UPLOAD_DIR = None

app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "*",
        "http://localhost:3000",
        "http://localhost:5000",
        "http://localhost:8000",
        "http://localhost:8080",
        "http://localhost:52000",
        "http://localhost:53000",
        "http://localhost:54000",
        "http://127.0.0.1:8000",
    ],
    allow_origin_regex=r"https://.*\.ngrok.*|http://localhost:.*|http://127\.0\.0\.1:.*",
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

oauth2_scheme = OAuth2PasswordBearer(tokenUrl="auth/login")


class SendOtpRequest(BaseModel):
    email: str


class VerifyOtpRequest(BaseModel):
    email: str
    otp: str
    full_name: str
    role: str
    phone: Optional[str] = None
    department: Optional[str] = None
    graduation_year: Optional[int] = None
    city: Optional[str] = None
    bio: Optional[str] = None
    company: Optional[str] = None
    job_title: Optional[str] = None
    mentorship_available: bool = True
    skills: Optional[str] = None
    interests: Optional[str] = None
    resume_url: Optional[str] = None


class InternshipCreate(BaseModel):
    posted_by: int
    role_title: str
    company: str
    location: Optional[str] = None
    duration: Optional[str] = None
    stipend: Optional[str] = None
    required_skills: Optional[str] = None
    seats: int = 1
    deadline: Optional[str] = None
    description: Optional[str] = None
    status: Optional[str] = "Open"


class ApplicationCreate(BaseModel):
    internship_id: int
    student_id: int
    cover_note: Optional[str] = None
    resume_url: Optional[str] = None


class ApplicationStatusUpdate(BaseModel):
    status: str


class EventCreate(BaseModel):
    created_by: int
    title: str
    date: Optional[str] = None
    location: Optional[str] = None
    description: Optional[str] = None
    category: str = "event"
    target_audience: str = "all"  # "all", "student", "alumni"
    has_document: bool = False
    has_photos: bool = False
    document_url: Optional[str] = None
    photo_url: Optional[str] = None


class AnnouncementCreate(BaseModel):
    created_by: int
    title: str
    content: str


class ConnectionCreate(BaseModel):
    requester_id: int
    receiver_id: int


class ConnectionStatusUpdate(BaseModel):
    status: str


class ResourceCreate(BaseModel):
    created_by: int
    title: str
    description: Optional[str] = None
    resource_type: str  # 'document', 'link'
    url: str
    target_audience: str = "all"  # "all", "student", "alumni"


class NotificationCreate(BaseModel):
    user_id: int
    type: Optional[str] = None
    message: str


class MessageCreate(BaseModel):
    sender_id: int
    receiver_id: int
    content: str


class AuthRegisterRequest(BaseModel):
    email: str
    password: str
    full_name: str
    role: str


class AuthLoginRequest(BaseModel):
    email: str
    password: str
    role: str
    fcm_token: Optional[str] = None


class ResetPasswordRequest(BaseModel):
    email: str
    role: str
    otp: str
    new_password: str


class ProfileUpdateRequest(BaseModel):
    full_name: Optional[str] = None
    phone: Optional[str] = None
    department: Optional[str] = None
    graduation_year: Optional[int] = None
    city: Optional[str] = None
    bio: Optional[str] = None
    company: Optional[str] = None
    job_title: Optional[str] = None
    mentorship_available: Optional[bool] = None
    skills: Optional[str] = None
    interests: Optional[str] = None
    resume_url: Optional[str] = None
    experience_summary: Optional[str] = None
    profile_picture_url: Optional[str] = None
    current_status: Optional[str] = None
    designation: Optional[str] = None
    responsibilities: Optional[str] = None
    educational_details: Optional[str] = None
    skills: Optional[str] = None
    interests: Optional[str] = None
    resume_url: Optional[str] = None


def payload_dict(payload: BaseModel):
    return payload.model_dump() if hasattr(payload, "model_dump") else payload.dict()


def ensure_role(role: str, allowed_roles: set[str]):
    normalized = role.lower()
    if normalized not in allowed_roles:
        raise HTTPException(status_code=400, detail=f"Role must be one of: {', '.join(sorted(allowed_roles))}")
    return normalized


def get_db():
    yield from get_session()


def get_user_by_id(db: Session, user_id: int):
    user = db.query(models.User).filter(models.User.user_id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    return user


def get_internship_by_id(db: Session, internship_id: int):
    internship = db.query(models.Internship).filter(models.Internship.internship_id == internship_id).first()
    if not internship:
        raise HTTPException(status_code=404, detail="Internship not found")
    return internship


def get_application_by_id(db: Session, application_id: int):
    application = db.query(models.Application).filter(models.Application.application_id == application_id).first()
    if not application:
        raise HTTPException(status_code=404, detail="Application not found")
    return application


def get_announcement_by_id(db: Session, announcement_id: int):
    announcement = db.query(models.Announcement).filter(models.Announcement.announcement_id == announcement_id).first()
    if not announcement:
        raise HTTPException(status_code=404, detail="Announcement not found")
    return announcement


def get_notification_by_id(db: Session, noti_id: int):
    notification = db.query(models.Notification).filter(models.Notification.noti_id == noti_id).first()
    if not notification:
        raise HTTPException(status_code=404, detail="Notification not found")
    return notification


def get_connection_by_id(db: Session, connection_id: int):
    connection = db.query(models.Connection).filter(models.Connection.connection_id == connection_id).first()
    if not connection:
        raise HTTPException(status_code=404, detail="Connection not found")
    return connection


def get_current_user(token: str = Depends(oauth2_scheme), db: Session = Depends(get_session)):
    credentials_exception = HTTPException(
        status_code=status.HTTP_401_UNAUTHORIZED,
        detail="Could not validate credentials",
        headers={"WWW-Authenticate": "Bearer"},
    )
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        user_id: str = payload.get("sub")
        if user_id is None:
            raise credentials_exception
    except JWTError:
        raise credentials_exception
    user = db.query(models.User).filter(models.User.user_id == int(user_id)).first()
    if user is None:
        raise credentials_exception
    return user


def _notify_audience(db, target_role: str, sender_id: int, noti_type: str, message: str):
    """Create a Notification row for every user matching target_role, excluding the sender."""
    query = db.query(models.User)
    if target_role != "all":
        query = query.filter(models.User.role == target_role)
    else:
        # "all" means students + alumni (not admin/staff themselves)
        query = query.filter(models.User.role.in_(["student", "alumni"]))

    for user in query.all():
        if user.user_id == sender_id:
            continue
        db.add(models.Notification(
            user_id=user.user_id,
            type=noti_type,
            message=message,
            is_read=False,
        ))
    db.commit()


def serialize_user(user: models.User):    return {
        "user_id": user.user_id,
        "email": user.email,
        "full_name": user.full_name,
        "role": user.role,
        "phone": user.phone,
        "department": user.department,
        "graduation_year": user.graduation_year,
        "city": user.city,
        "bio": user.bio,
        "profile_picture_url": user.profile_picture_url,
        "current_status": user.current_status,
        "designation": user.designation,
        "responsibilities": user.responsibilities,
        "is_verified": user.is_verified,
    }


def serialize_alumni(user: models.User, profile: models.AlumniProfile):
    data = serialize_user(user)
    data.update(
        {
            "company": profile.company,
            "job_title": profile.job_title,
            "mentorship_available": profile.mentorship_available,
            "experience_summary": profile.experience_summary,
        }
    )
    return data


def serialize_student(user: models.User, profile: models.StudentProfile):
    data = serialize_user(user)
    data.update(
        {
            "skills": profile.educational_details,
            "interests": profile.interests,
            "resume_url": profile.resume_url,
            "educational_details": profile.educational_details,
        }
    )
    return data


def ensure_connected(db: Session, sender_id: int, receiver_id: int):
    if sender_id == receiver_id:
        return None

    sender = get_user_by_id(db, sender_id)
    receiver = get_user_by_id(db, receiver_id)

    # Only gate student <-> alumni chats
    # student <-> admin/staff is always free
    # staff can message anyone freely
    roles = {sender.role, receiver.role}
    if "staff" in roles or roles != {"student", "alumni"}:
        return None

    connection = db.query(models.Connection).filter(
        (
            (models.Connection.requester_id == sender_id)
            & (models.Connection.receiver_id == receiver_id)
        )
        | (
            (models.Connection.requester_id == receiver_id)
            & (models.Connection.receiver_id == sender_id)
        )
    ).filter(models.Connection.status == "accepted").first()

    if not connection:
        raise HTTPException(
            status_code=403,
            detail="Messaging is available only after the alumni accepts the connection request",
        )

    return None


@app.get("/")
def home():
    return {
        "message": "Alumni App backend is running",
        "modules": ["auth", "directory", "internships", "applications", "events"],
    }


@app.post("/send-otp", tags=["auth"])
def send_otp(payload: SendOtpRequest):
    send_otp_email(payload.email)
    return {"message": "OTP sent", "expires_in_minutes": 5}


@app.post("/auth/register", tags=["auth"])
def register_with_password(payload: AuthRegisterRequest, current_user: models.User = Depends(get_current_user), db: Session = Depends(get_db)):
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Only admins can register new users")
    
    role = ensure_role(payload.role, {"student", "alumni", "admin", "staff"})

    existing_user = db.query(models.User).filter(models.User.email == payload.email).first()
    if existing_user:
        raise HTTPException(status_code=400, detail="User already exists with this email")

    user = models.User(
        email=payload.email,
        full_name=payload.full_name,
        role=role,
        password_hash=hash_password(payload.password),
        is_verified=True,
    )
    db.add(user)
    db.flush()

    if role == "student":
        db.add(models.StudentProfile(student_id=user.user_id))
    if role == "alumni":
        db.add(models.AlumniProfile(alumni_id=user.user_id))

    db.commit()
    db.refresh(user)

    return {
        "status": "registered",
        "user_id": user.user_id,
        "email": user.email,
        "full_name": user.full_name,
        "role": user.role,
    }


@app.post("/auth/bulk-register", tags=["auth"])
async def bulk_register(
    file: UploadFile = File(...),
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """
    Upload a CSV or Excel file to bulk register students/alumni.
    Required columns: full_name, email, password, role
    Optional columns: phone, department, graduation_year, city, company, job_title
    """
    import io
    import csv

    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Only admins can bulk register users")

    content = await file.read()
    filename = file.filename.lower()

    rows = []

    if filename.endswith(".csv"):
        decoded = content.decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(decoded))
        rows = list(reader)
    elif filename.endswith(".xlsx") or filename.endswith(".xls"):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(content))
            ws = wb.active
            headers = [cell.value for cell in ws[1]]
            for row in ws.iter_rows(min_row=2, values_only=True):
                rows.append(dict(zip(headers, row)))
        except ImportError:
            raise HTTPException(status_code=400, detail="openpyxl not installed. Use CSV format instead.")
    else:
        raise HTTPException(status_code=400, detail="Only .csv, .xlsx, or .xls files are supported")

    required_cols = {"full_name", "email", "password", "role"}
    if rows and not required_cols.issubset(set(rows[0].keys())):
        raise HTTPException(
            status_code=400,
            detail=f"Missing required columns. Need: {', '.join(required_cols)}"
        )

    results = {"success": [], "skipped": [], "errors": []}

    for i, row in enumerate(rows, start=2):
        try:
            email = str(row.get("email", "")).strip()
            full_name = str(row.get("full_name", "")).strip()
            password = str(row.get("password", "")).strip()
            role = str(row.get("role", "")).strip().lower()

            if not email or not full_name or not password or not role:
                results["errors"].append({"row": i, "reason": "Missing required field"})
                continue

            if role not in {"student", "alumni", "staff"}:
                results["errors"].append({"row": i, "email": email, "reason": f"Invalid role '{role}'"})
                continue

            existing = db.query(models.User).filter(models.User.email == email).first()
            if existing:
                results["skipped"].append({"row": i, "email": email, "reason": "Email already exists"})
                continue

            user = models.User(
                email=email,
                full_name=full_name,
                role=role,
                password_hash=hash_password(password),
                is_verified=True,
                phone=str(row.get("phone", "") or "").strip() or None,
                department=str(row.get("department", "") or "").strip() or None,
                graduation_year=int(row["graduation_year"]) if row.get("graduation_year") else None,
                city=str(row.get("city", "") or "").strip() or None,
            )
            db.add(user)
            db.flush()

            if role == "student":
                db.add(models.StudentProfile(student_id=user.user_id))
            if role == "alumni":
                profile = models.AlumniProfile(alumni_id=user.user_id)
                profile.company = str(row.get("company", "") or "").strip() or None
                profile.job_title = str(row.get("job_title", "") or "").strip() or None
                db.add(profile)

            db.commit()
            results["success"].append({"row": i, "email": email, "role": role})

        except Exception as e:
            db.rollback()
            results["errors"].append({"row": i, "reason": str(e)})

    return {
        "total": len(rows),
        "registered": len(results["success"]),
        "skipped": len(results["skipped"]),
        "errors": len(results["errors"]),
        "details": results,
    }


@app.post("/auth/login", tags=["auth"])
def login_with_password(payload: AuthLoginRequest, db: Session = Depends(get_db)):
    role = ensure_role(payload.role, {"student", "alumni", "admin", "staff"})
    user = db.query(models.User).filter(
        models.User.email == payload.email,
        models.User.role == role,
    ).first()

    print(payload)

    if not user:
        print(f"Login failed: User '{payload.email}' with role '{role}' not found.")
        raise HTTPException(status_code=401, detail="Invalid email, role, or password")

    if not verify_password(payload.password, user.password_hash):
        print(f"Login failed: Incorrect password for '{payload.email}'.")
        raise HTTPException(status_code=401, detail="Invalid email, role, or password")

    # Store FCM token if provided
    if payload.fcm_token:
        user.fcm_token = payload.fcm_token
        db.commit()

    access_token = create_access_token(data={"sub": str(user.user_id), "role": user.role})

    return {
        "status": "authenticated",
        "access_token": access_token,
        "token_type": "bearer",
        "user_id": user.user_id,
        "email": user.email,
        "full_name": user.full_name,
        "role": user.role,
    }


class FcmTokenRequest(BaseModel):
    user_id: int
    token: str

@app.post("/auth/fcm-token", tags=["auth"])
def save_fcm_token(
    payload: FcmTokenRequest,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    # Use authenticated user — ignore user_id in body for security
    current_user.fcm_token = payload.token
    db.commit()
    return {"status": "token_saved"}


@app.post("/auth/test-fcm", tags=["auth"])
def test_fcm_notification(db: Session = Depends(get_db)):
    """Send a test FCM push to all admin devices. No auth required — for testing only."""
    admins = db.query(models.User).filter(
        models.User.role == "admin",
        models.User.fcm_token.isnot(None),
    ).all()

    if not admins:
        return {"status": "no_targets", "message": "No admin devices with FCM token found"}

    import threading
    for admin in admins:
        threading.Thread(
            target=_send_fcm_push,
            args=(
                admin.fcm_token,
                "Alumni Network",
                "FCM is working! Push notifications are set up correctly.",
            ),
            daemon=True,
        ).start()

    return {"status": "sent", "recipients": len(admins)}


@app.post("/auth/reset-password", tags=["auth"])
def reset_password(payload: ResetPasswordRequest, db: Session = Depends(get_db)):
    if not verify_stored_otp(payload.email, payload.otp):
        raise HTTPException(status_code=400, detail="Invalid or expired OTP")
    
    role = ensure_role(payload.role, {"student", "alumni", "admin", "staff"})
    user = db.query(models.User).filter(
        models.User.email == payload.email,
        models.User.role == role,
    ).first()

    if not user:
        raise HTTPException(status_code=404, detail="User not found")

    user.password_hash = hash_password(payload.new_password)
    db.commit()
    return {"status": "password_updated"}


@app.post("/verify-otp", tags=["auth"])
def verify_otp(payload: VerifyOtpRequest, db: Session = Depends(get_db)):
    if not verify_stored_otp(payload.email, payload.otp):
        raise HTTPException(status_code=400, detail="Invalid OTP")

    role = ensure_role(payload.role, {"student", "alumni", "admin", "staff"})

    user = db.query(models.User).filter(models.User.email == payload.email).first()
    if not user:
        user = models.User(email=payload.email, full_name=payload.full_name, role=role)
        db.add(user)
        db.flush()

    user.full_name = payload.full_name
    user.role = role
    user.phone = payload.phone
    user.department = payload.department
    user.graduation_year = payload.graduation_year
    user.city = payload.city
    user.bio = payload.bio
    user.is_verified = True

    if role == "student":
        profile = db.query(models.StudentProfile).filter(models.StudentProfile.student_id == user.user_id).first()
        if not profile:
            profile = models.StudentProfile(student_id=user.user_id)
            db.add(profile)
        profile.skills = payload.skills
        profile.interests = payload.interests
        profile.resume_url = payload.resume_url

    if role == "alumni":
        profile = db.query(models.AlumniProfile).filter(models.AlumniProfile.alumni_id == user.user_id).first()
        if not profile:
            profile = models.AlumniProfile(alumni_id=user.user_id)
            db.add(profile)
        profile.company = payload.company
        profile.job_title = payload.job_title
        profile.mentorship_available = payload.mentorship_available
        profile.experience_summary = payload.bio

    db.commit()
    db.refresh(user)
    clear_otp(payload.email)

    access_token = create_access_token(data={"sub": str(user.user_id), "role": user.role})

    return {
        "status": "verified",
        "access_token": access_token,
        "token_type": "bearer",
        "user_id": user.user_id,
        "role": user.role,
    }


@app.get("/alumni", tags=["alumni"])
def list_alumni(
    department: Optional[str] = None,
    company: Optional[str] = None,
    city: Optional[str] = None,
    job_title: Optional[str] = None,
    db: Session = Depends(get_db),
):
    query = db.query(models.User, models.AlumniProfile).join(
        models.AlumniProfile, models.AlumniProfile.alumni_id == models.User.user_id
    )

    if department:
        query = query.filter(models.User.department.ilike(f"%{department}%"))
    if company:
        query = query.filter(models.AlumniProfile.company.ilike(f"%{company}%"))
    if city:
        query = query.filter(models.User.city.ilike(f"%{city}%"))
    if job_title:
        query = query.filter(models.AlumniProfile.job_title.ilike(f"%{job_title}%"))

    results = query.all()
    return [serialize_alumni(user, profile) for user, profile in results]


# ── Student Streak ────────────────────────────────────────────────────────────

@app.post("/student/checkin", tags=["profile"])
def student_checkin(
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    if current_user.role != "student":
        raise HTTPException(status_code=403, detail="Students only")

    import json as _json
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")

    row = db.query(models.StudentStreak).filter(
        models.StudentStreak.student_id == current_user.user_id
    ).first()

    if not row:
        row = models.StudentStreak(
            student_id=current_user.user_id,
            current_streak=1,
            longest_streak=1,
            last_checkin=today,
            checkin_history=_json.dumps([today]),
        )
        db.add(row)
        db.commit()
        return {"streak": 1, "longest": 1, "already_checked_in": False}

    if row.last_checkin == today:
        history = _json.loads(row.checkin_history or "[]")
        return {"streak": row.current_streak, "longest": row.longest_streak, "already_checked_in": True, "history": history[-30:]}

    history = _json.loads(row.checkin_history or "[]")
    history.append(today)
    history = history[-90:]  # keep last 90 days

    # Calculate new streak
    if row.last_checkin:
        from datetime import date as _date
        last = _date.fromisoformat(row.last_checkin)
        tod = _date.fromisoformat(today)
        diff = (tod - last).days
        new_streak = row.current_streak + 1 if diff == 1 else 1
    else:
        new_streak = 1

    new_longest = max(row.longest_streak, new_streak)
    row.current_streak = new_streak
    row.longest_streak = new_longest
    row.last_checkin = today
    row.checkin_history = _json.dumps(history)
    db.commit()

    return {"streak": new_streak, "longest": new_longest, "already_checked_in": False, "history": history[-30:]}


@app.get("/student/streak", tags=["profile"])
def get_streak(
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    import json as _json
    row = db.query(models.StudentStreak).filter(
        models.StudentStreak.student_id == current_user.user_id
    ).first()
    if not row:
        return {"streak": 0, "longest": 0, "last_checkin": None, "history": []}
    history = _json.loads(row.checkin_history or "[]")
    return {
        "streak": row.current_streak,
        "longest": row.longest_streak,
        "last_checkin": row.last_checkin,
        "history": history[-30:],
    }


@app.get("/student/dashboard", tags=["profile"])
def get_student_dashboard(
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    if current_user.role != "student":
        raise HTTPException(status_code=403, detail="Students only")

    uid = current_user.user_id

    # Accepted connections count
    from sqlalchemy import or_
    connections = db.query(models.Connection).filter(
        models.Connection.status == "accepted",
        or_(
            models.Connection.requester_id == uid,
            models.Connection.receiver_id == uid,
        )
    ).count()

    # Unread messages
    unread_msgs = db.query(models.Message).filter(
        models.Message.receiver_id == uid,
        models.Message.is_read == False,
    ).count()

    # Applications sent
    apps_sent = db.query(models.Application).filter(
        models.Application.student_id == uid
    ).count()

    print(f"[Dashboard] uid={uid} connections={connections} unread={unread_msgs} apps={apps_sent}")

    # Latest 3 events (notice board)
    events = db.query(models.Event).filter(
        (models.Event.target_audience == "all") | (models.Event.target_audience == "student")
    ).order_by(models.Event.created_at.desc()).limit(5).all()

    # Featured internship (latest open)
    featured = db.query(models.Internship).filter(
        models.Internship.status == "Open"
    ).order_by(models.Internship.created_at.desc()).first()

    # Recent activity from notifications
    recent = db.query(models.Notification).filter(
        models.Notification.user_id == uid
    ).order_by(models.Notification.created_at.desc()).limit(5).all()

    # Find a mentor — alumni with mentorship available
    mentors = db.query(models.User, models.AlumniProfile).join(
        models.AlumniProfile, models.AlumniProfile.alumni_id == models.User.user_id
    ).filter(
        models.AlumniProfile.mentorship_available == True
    ).limit(3).all()

    return {
        "stats": {
            "connections": connections,
            "unread_messages": unread_msgs,
            "applications_sent": apps_sent,
        },
        "notices": [
            {
                "event_id": e.event_id,
                "title": e.title,
                "date": e.date,
                "description": (e.description or "")[:120],
                "category": e.category,
            } for e in events
        ],
        "featured_internship": {
            "internship_id": featured.internship_id,
            "role_title": featured.role_title,
            "company": featured.company,
            "deadline": featured.deadline,
            "location": featured.location,
        } if featured else None,
        "recent_activity": [
            {
                "message": r.message,
                "type": r.type,
                "created_at": r.created_at.isoformat() if r.created_at else None,
                "is_read": r.is_read,
            } for r in recent
        ],
        "mentors": [
            {
                "user_id": u.user_id,
                "full_name": u.full_name,
                "job_title": p.job_title,
                "company": p.company,
                "profile_picture_url": u.profile_picture_url,
            } for u, p in mentors
        ],
    }


@app.get("/admin/stats", tags=["admin"])
def get_admin_stats(
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    if current_user.role not in {"admin", "staff"}:
        raise HTTPException(status_code=403, detail="Not authorized")

    total_students  = db.query(models.User).filter(models.User.role == "student").count()
    total_alumni    = db.query(models.User).filter(models.User.role == "alumni").count()
    total_staff     = db.query(models.User).filter(models.User.role == "staff").count()
    total_admins    = db.query(models.User).filter(models.User.role == "admin").count()
    total_internships = db.query(models.Internship).count()
    total_events    = db.query(models.Event).count()
    total_resources = db.query(models.Resource).count()
    total_connections = db.query(models.Connection).filter(models.Connection.status == "accepted").count()
    pending_connections = db.query(models.Connection).filter(models.Connection.status == "pending").count()
    total_messages  = db.query(models.Message).count()

    return {
        "total_students": total_students,
        "total_alumni": total_alumni,
        "total_staff": total_staff,
        "total_admins": total_admins,
        "total_internships": total_internships,
        "total_events": total_events,
        "total_resources": total_resources,
        "total_connections": total_connections,
        "pending_connections": pending_connections,
        "total_messages": total_messages,
        "total_users": total_students + total_alumni + total_staff + total_admins,
    }


@app.get("/users", tags=["admin"])
def list_users(
    role: Optional[str] = None,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    if current_user.role not in {"admin", "staff"}:
        raise HTTPException(status_code=403, detail="Only admins or staff can list all users")
    
    query = db.query(models.User)
    if role:
        query = query.filter(models.User.role == role)
    
    users = query.all()
    return [serialize_user(u) for u in users]


@app.get("/admins", tags=["admin"])
def list_admin_contacts(
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    if current_user.role not in {"admin", "alumni", "student", "staff"}:
        raise HTTPException(status_code=403, detail="Not authorized to view admin contacts")

    admins = (
        db.query(models.User)
        .filter(models.User.role == "admin")
        .order_by(models.User.full_name.asc())
        .all()
    )
    return [serialize_user(admin) for admin in admins]


@app.get("/staff-contacts", tags=["admin"])
def list_staff_contacts(
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Returns all staff users — accessible to students for direct messaging."""
    staff = (
        db.query(models.User)
        .filter(models.User.role == "staff")
        .order_by(models.User.full_name.asc())
        .all()
    )
    return [serialize_user(s) for s in staff]


@app.get("/profile/{user_id}", tags=["profile"])
def get_profile(user_id: int, db: Session = Depends(get_db)):
    user = get_user_by_id(db, user_id)
    data = serialize_user(user)

    if user.role == "student":
        profile = db.query(models.StudentProfile).filter(models.StudentProfile.student_id == user.user_id).first()
        data.update(
            {
                "skills": profile.skills if profile else None,
                "interests": profile.interests if profile else None,
                "resume_url": profile.resume_url if profile else None,
                "educational_details": profile.educational_details if profile else None,
            }
        )

    if user.role == "alumni":
        profile = db.query(models.AlumniProfile).filter(models.AlumniProfile.alumni_id == user.user_id).first()
        data.update(
            {
                "company": profile.company if profile else None,
                "job_title": profile.job_title if profile else None,
                "mentorship_available": profile.mentorship_available if profile else True,
                "experience_summary": profile.experience_summary if profile else None,
            }
        )

    return data


@app.get("/auth/me", tags=["auth"])
def read_users_me(current_user: models.User = Depends(get_current_user), db: Session = Depends(get_db)):
    return get_profile(current_user.user_id, db)


@app.patch("/profile/{user_id}", tags=["profile"])
def update_profile(user_id: int, payload: ProfileUpdateRequest, db: Session = Depends(get_db)):
    user = get_user_by_id(db, user_id)
    updates = payload.model_dump(exclude_unset=True) if hasattr(payload, "model_dump") else payload.dict(exclude_unset=True)

    for field, value in updates.items():
        if field in ["full_name", "phone", "department", "graduation_year", "city", "bio", "profile_picture_url", "current_status", "designation", "responsibilities"]:
            setattr(user, field, value)

    if user.role == "student":
        profile = db.query(models.StudentProfile).filter(models.StudentProfile.student_id == user.user_id).first()
        if not profile:
            profile = models.StudentProfile(student_id=user.user_id)
            db.add(profile)
        for field in ["skills", "interests", "resume_url", "educational_details"]:
            value = updates.get(field)
            if value is not None:
                setattr(profile, field, value)

    if user.role == "alumni":
        profile = db.query(models.AlumniProfile).filter(models.AlumniProfile.alumni_id == user.user_id).first()
        if not profile:
            profile = models.AlumniProfile(alumni_id=user.user_id)
            db.add(profile)
        if updates.get("company") is not None:
            profile.company = updates["company"]
        if updates.get("job_title") is not None:
            profile.job_title = updates["job_title"]
        if updates.get("mentorship_available") is not None:
            profile.mentorship_available = updates["mentorship_available"]
        if updates.get("experience_summary") is not None:
            profile.experience_summary = updates["experience_summary"]

    db.commit()
    return {"status": "updated"}


class EmailChangeRequest(BaseModel):
    user_id: int
    new_email: str


class EmailChangeVerifyRequest(BaseModel):
    user_id: int
    new_email: str
    otp: str


@app.post("/profile/request-email-change", tags=["profile"])
def request_email_change(payload: EmailChangeRequest, db: Session = Depends(get_db)):
    user = get_user_by_id(db, payload.user_id)

    # Check new email not already taken
    existing = db.query(models.User).filter(
        models.User.email == payload.new_email,
        models.User.user_id != payload.user_id,
    ).first()
    if existing:
        raise HTTPException(status_code=400, detail="Email already in use by another account")

    send_otp_email_change(payload.new_email)
    return {"message": f"OTP sent to {payload.new_email}"}


@app.post("/profile/verify-email-change", tags=["profile"])
def verify_email_change(payload: EmailChangeVerifyRequest, db: Session = Depends(get_db)):
    if not verify_stored_otp(payload.new_email, payload.otp):
        raise HTTPException(status_code=400, detail="Invalid or expired OTP")

    user = get_user_by_id(db, payload.user_id)
    user.email = payload.new_email
    db.commit()
    clear_otp(payload.new_email)
    return {"status": "email_updated", "new_email": payload.new_email}


@app.delete("/profile/{user_id}")
def delete_profile(user_id: int, current_user: models.User = Depends(get_current_user), db: Session = Depends(get_db)):
    if current_user.user_id != user_id and current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Not authorized to delete this profile")
    
    user = get_user_by_id(db, user_id)
    try:
        # Remove dependent records first because most foreign keys do not use CASCADE.
        db.query(models.Connection).filter(
            (models.Connection.requester_id == user_id)
            | (models.Connection.receiver_id == user_id)
        ).delete(synchronize_session=False)

        db.query(models.Message).filter(
            (models.Message.sender_id == user_id)
            | (models.Message.receiver_id == user_id)
        ).delete(synchronize_session=False)

        db.query(models.Notification).filter(
            models.Notification.user_id == user_id
        ).delete(synchronize_session=False)

        db.query(models.Application).filter(
            models.Application.student_id == user_id
        ).delete(synchronize_session=False)

        internship_ids = [
            internship_id
            for (internship_id,) in db.query(models.Internship.internship_id)
            .filter(models.Internship.posted_by == user_id)
            .all()
        ]
        if internship_ids:
            db.query(models.Application).filter(
                models.Application.internship_id.in_(internship_ids)
            ).delete(synchronize_session=False)
            db.query(models.Internship).filter(
                models.Internship.internship_id.in_(internship_ids)
            ).delete(synchronize_session=False)

        db.query(models.Event).filter(
            models.Event.created_by == user_id
        ).delete(synchronize_session=False)

        db.query(models.Announcement).filter(
            models.Announcement.created_by == user_id
        ).delete(synchronize_session=False)

        db.query(models.StudentProfile).filter(
            models.StudentProfile.student_id == user_id
        ).delete(synchronize_session=False)
        db.query(models.AlumniProfile).filter(
            models.AlumniProfile.alumni_id == user_id
        ).delete(synchronize_session=False)

        db.delete(user)
        db.commit()
    except Exception:
        db.rollback()
        raise

    return {"status": "deleted"}


@app.post("/profile/{user_id}/resume", tags=["profile"])
def upload_resume(user_id: int, file: UploadFile = File(...), db: Session = Depends(get_db)):
    user = get_user_by_id(db, user_id)

    if user.role != "student":
        raise HTTPException(status_code=400, detail="Resume upload is only available for student users")

    filename = f"user_{user_id}_{file.filename}".replace(" ", "_")
    file_path = os.path.join(UPLOAD_DIR, filename)
    with open(file_path, "wb") as buffer:
        shutil.copyfileobj(file.file, buffer)

    resume_url = f"/uploads/{filename}"
    profile = db.query(models.StudentProfile).filter(models.StudentProfile.student_id == user.user_id).first()
    if not profile:
        profile = models.StudentProfile(student_id=user.user_id)
        db.add(profile)
    profile.resume_url = resume_url
    db.commit()

    return {"status": "uploaded", "resume_url": resume_url}


@app.post("/upload", tags=["common"])
def upload_file(file: UploadFile = File(...)):
    cloudinary_url = os.getenv("CLOUDINARY_URL")

    if cloudinary_url:
        # Upload to Cloudinary — permanent, survives redeploys
        import cloudinary
        import cloudinary.uploader
        cloudinary.config(cloudinary_url=cloudinary_url)
        contents = file.file.read()
        result = cloudinary.uploader.upload(
            contents,
            folder="alumni_network",
            resource_type="auto",
        )
        print(result["secure_url"])
        return {"status": "uploaded", "url": result["secure_url"]}

    # Fallback: local disk (dev only — not persistent on Render)
    import time
    ext = os.path.splitext(file.filename)[1]
    filename = f"common_{int(time.time() * 1000)}{ext}".replace(" ", "_")
    file_path = os.path.join(UPLOAD_DIR, filename)
    with open(file_path, "wb") as buffer:
        shutil.copyfileobj(file.file, buffer)
    return {"status": "uploaded", "url": f"/uploads/{filename}"}


@app.post("/internships", tags=["internships"])
def create_internship(payload: InternshipCreate, db: Session = Depends(get_db)):
    user = get_user_by_id(db, payload.posted_by)
    if user.role != "alumni":
        raise HTTPException(status_code=400, detail="Only alumni users can post internships")

    internship = models.Internship(**payload_dict(payload))
    db.add(internship)
    db.commit()
    db.refresh(internship)

    # Award points for posting internship
    _award_points(db, user.user_id, "internship", 10)
    db.commit()

    notif_title = f"💼 New Internship: {payload.role_title}"
    notif_body = f"{payload.company}" + (f" | {payload.location}" if payload.location else "") + f" — Posted by {user.full_name}"

    # Get connected students of this alumni
    connections = db.query(models.Connection).filter(
        models.Connection.requester_id == user.user_id,
        models.Connection.status == "accepted",
    ).all()
    connected_student_ids = {c.receiver_id for c in connections}

    # Also get connections where alumni is receiver
    connections2 = db.query(models.Connection).filter(
        models.Connection.receiver_id == user.user_id,
        models.Connection.status == "accepted",
    ).all()
    connected_student_ids.update(c.requester_id for c in connections2)

    # Get admin and staff users
    staff_admins = db.query(models.User).filter(
        models.User.role.in_(["admin", "staff"])
    ).all()
    staff_admin_ids = {u.user_id for u in staff_admins}

    # All target user IDs
    target_ids = connected_student_ids | staff_admin_ids

    import threading
    for target_id in target_ids:
        # Save to notification DB
        db.add(models.Notification(
            user_id=target_id,
            type="event",
            message=f"{notif_title} — {notif_body}",
            is_read=False,
        ))

    db.commit()

    # Send FCM push to targets with tokens
    targets_with_tokens = db.query(models.User).filter(
        models.User.user_id.in_(target_ids),
        models.User.fcm_token.isnot(None),
    ).all()

    for target in targets_with_tokens:
        threading.Thread(
            target=_send_fcm_push,
            args=(target.fcm_token, notif_title, notif_body),
            daemon=True,
        ).start()

    return {"internship_id": internship.internship_id, "message": "Internship created"}


@app.get("/internships", tags=["internships"])
def list_internships(
    role_title: Optional[str] = None,
    location: Optional[str] = None,
    stipend: Optional[str] = None,
    db: Session = Depends(get_db),
):
    query = db.query(models.Internship)
    if role_title:
        query = query.filter(models.Internship.role_title.ilike(f"%{role_title}%"))
    if location:
        query = query.filter(models.Internship.location.ilike(f"%{location}%"))
    if stipend:
        query = query.filter(models.Internship.stipend.ilike(f"%{stipend}%"))
    internships = query.order_by(models.Internship.created_at.desc()).all()

    result = []
    for i in internships:
        poster = db.query(models.User).filter(models.User.user_id == i.posted_by).first()
        result.append({
            "internship_id": i.internship_id,
            "posted_by": i.posted_by,
            "posted_by_name": poster.full_name if poster else "Unknown",
            "company_name": i.company,
            "role_title": i.role_title,
            "description": i.description,
            "location": i.location,
            "duration": i.duration,
            "stipend": i.stipend,
            "skills_required": i.required_skills,
            "apply_deadline": i.deadline,
            "seats_available": i.seats,
            "status": i.status,
            "created_at": i.created_at.isoformat(),
        })
    return result


@app.post("/applications", tags=["applications"])
def apply_for_internship(payload: ApplicationCreate, db: Session = Depends(get_db)):
    student = get_user_by_id(db, payload.student_id)
    internship = get_internship_by_id(db, payload.internship_id)

    if student.role != "student":
        raise HTTPException(status_code=400, detail="Only students can apply for internships")

    existing = db.query(models.Application).filter(
        models.Application.internship_id == payload.internship_id,
        models.Application.student_id == payload.student_id,
    ).first()
    if existing:
        raise HTTPException(status_code=400, detail="Student already applied for this internship")

    application = models.Application(**payload_dict(payload))
    db.add(application)
    db.flush()

    # Notify the alumni who posted this internship
    notification = models.Notification(
        user_id=internship.posted_by,
        message=f"{student.full_name} applied for your '{internship.role_title}' internship at {internship.company}.",
        is_read=False,
    )
    db.add(notification)
    db.commit()
    db.refresh(application)
    return {"application_id": application.application_id, "status": application.status}


@app.get("/internships/{internship_id}/applicants", tags=["applications"])
def get_internship_applicants(
    internship_id: int,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    internship = get_internship_by_id(db, internship_id)
    if current_user.role not in {"alumni", "admin"}:
        raise HTTPException(status_code=403, detail="Only alumni or admin can view applicants")
    if current_user.role == "alumni" and internship.posted_by != current_user.user_id:
        raise HTTPException(status_code=403, detail="You can only view applicants for your own internships")

    applications = db.query(models.Application).filter(
        models.Application.internship_id == internship_id
    ).order_by(models.Application.applied_at.desc()).all()

    result = []
    for app in applications:
        student = db.query(models.User).filter(models.User.user_id == app.student_id).first()
        student_profile = db.query(models.StudentProfile).filter(
            models.StudentProfile.student_id == app.student_id
        ).first()
        result.append({
            "application_id": app.application_id,
            "status": app.status,
            "cover_note": app.cover_note,
            "applied_at": app.applied_at.isoformat(),
            "student": {
                "user_id": student.user_id,
                "full_name": student.full_name,
                "email": student.email,
                "phone": student.phone,
                "department": student.department,
                "graduation_year": student.graduation_year,
                "city": student.city,
                "skills": student_profile.skills if student_profile else None,
                "resume_url": student_profile.resume_url if student_profile else None,
            } if student else None,
        })
    return result


@app.patch("/applications/{application_id}", tags=["applications"])
def update_application_status(
    application_id: int,
    payload: ApplicationStatusUpdate,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    if current_user.role != "alumni":
        raise HTTPException(status_code=403, detail="Only alumni can update application status")

    allowed_statuses = {"Applied", "Shortlisted", "Rejected", "Selected"}
    if payload.status not in allowed_statuses:
        raise HTTPException(status_code=400, detail="Invalid application status")

    application = get_application_by_id(db, application_id)

    # Ensure the alumni owns the internship this application belongs to
    internship = get_internship_by_id(db, application.internship_id)
    if internship.posted_by != current_user.user_id:
        raise HTTPException(status_code=403, detail="You can only update status for your own internship applicants")

    application.status = payload.status
    db.commit()
    return {"application_id": application.application_id, "status": application.status}


@app.patch("/internships/{internship_id}", tags=["internships"])
def update_internship(internship_id: int, payload: InternshipCreate, db: Session = Depends(get_db)):
    internship = get_internship_by_id(db, internship_id)
    updates = payload_dict(payload)
    for field, value in updates.items():
        if value is not None:
            setattr(internship, field, value)
    db.commit()
    return {"status": "updated"}


@app.delete("/internships/{internship_id}", tags=["internships"])
def delete_internship(internship_id: int, db: Session = Depends(get_db)):
    internship = get_internship_by_id(db, internship_id)
    # Delete related applications first to avoid foreign key constraint errors
    db.query(models.Application).filter(
        models.Application.internship_id == internship_id
    ).delete(synchronize_session=False)
    db.delete(internship)
    db.commit()
    return {"status": "deleted"}


@app.delete("/applications/{application_id}", tags=["applications"])
def delete_application(application_id: int, db: Session = Depends(get_db)):
    application = get_application_by_id(db, application_id)
    db.delete(application)
    db.commit()
    return {"status": "deleted"}


@app.get("/applications", tags=["applications"])
def list_applications(
    internship_id: Optional[int] = None,
    student_id: Optional[int] = None,
    db: Session = Depends(get_db),
):
    query = db.query(models.Application)
    if internship_id:
        query = query.filter(models.Application.internship_id == internship_id)
    if student_id:
        query = query.filter(models.Application.student_id == student_id)
    return query.order_by(models.Application.applied_at.desc()).all()


@app.post("/events", tags=["events"])
def create_event(payload: EventCreate, db: Session = Depends(get_db)):
    user = get_user_by_id(db, payload.created_by)
    if user.role not in {"admin", "staff"}:
        raise HTTPException(status_code=400, detail="Only admin or staff can create events")

    event = models.Event(**payload_dict(payload))
    db.add(event)
    db.commit()
    db.refresh(event)

    notif_title = f"📅 New Event: {payload.title}"
    notif_body = (f"{payload.date}" if payload.date else "") + \
                 (f" | {payload.location}" if payload.location else "") or \
                 (payload.description[:80] if payload.description else "Check it out!")

    # Save to notification DB
    _notify_audience(
        db=db,
        target_role=payload.target_audience,
        sender_id=user.user_id,
        noti_type="event",
        message=f"{notif_title} — {notif_body}",
    )

    # Send FCM push to target users with tokens
    import threading
    query = db.query(models.User).filter(models.User.fcm_token.isnot(None))
    if payload.target_audience != "all":
        query = query.filter(models.User.role == payload.target_audience)
    else:
        query = query.filter(models.User.role.in_(["student", "alumni"]))

    for target in query.all():
        if target.user_id == user.user_id:
            continue
        threading.Thread(
            target=_send_fcm_push,
            args=(target.fcm_token, notif_title, notif_body),
            daemon=True,
        ).start()

    return {"event_id": event.event_id, "message": "Event created"}


@app.get("/events", tags=["events"])
def list_events(current_user: models.User = Depends(get_current_user), db: Session = Depends(get_db)):
    if current_user.role in {"admin", "staff"}:
        return db.query(models.Event).order_by(models.Event.created_at.desc()).all()
    
    return db.query(models.Event).filter(
        (models.Event.target_audience == "all") | (models.Event.target_audience == current_user.role)
    ).order_by(models.Event.created_at.desc()).all()


@app.patch("/events/{event_id}", tags=["events"])
def update_event(event_id: int, payload: EventCreate, current_user: models.User = Depends(get_current_user), db: Session = Depends(get_db)):
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Only admins can update events")
    
    event = db.query(models.Event).filter(models.Event.event_id == event_id).first()
    if not event:
        raise HTTPException(status_code=404, detail="Event not found")

    event.title = payload.title
    event.date = payload.date
    event.location = payload.location
    event.description = payload.description
    event.category = payload.category
    event.target_audience = payload.target_audience
    db.commit()
    return {"status": "updated", "event_id": event.event_id}


@app.delete("/events/{event_id}", tags=["events"])
def delete_event(event_id: int, current_user: models.User = Depends(get_current_user), db: Session = Depends(get_db)):
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Only admins can delete events")

    event = db.query(models.Event).filter(models.Event.event_id == event_id).first()
    if not event:
        raise HTTPException(status_code=404, detail="Event not found")

    db.delete(event)
    db.commit()
    return {"status": "deleted"}


@app.delete("/events", tags=["events"])
def delete_all_events(current_user: models.User = Depends(get_current_user), db: Session = Depends(get_db)):
    if current_user.role not in {"admin", "staff"}:
        raise HTTPException(status_code=403, detail="Only admins or staff can delete all events")

    db.query(models.Event).delete()
    db.commit()
    return {"status": "all deleted"}


@app.post("/connections", tags=["connections"])
def create_connection(
    payload: ConnectionCreate,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    requester = get_user_by_id(db, payload.requester_id)
    receiver = get_user_by_id(db, payload.receiver_id)

    if current_user.user_id != payload.requester_id and current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Not authorized to send this connection request")
    if requester.user_id == receiver.user_id:
        raise HTTPException(status_code=400, detail="Users cannot connect to themselves")
    if requester.role != "student":
        raise HTTPException(status_code=400, detail="Only students can initiate connection requests")
    if receiver.role != "alumni":
        raise HTTPException(status_code=400, detail="Students can only connect with alumni")

    existing = db.query(models.Connection).filter(
        (
            (models.Connection.requester_id == payload.requester_id)
            & (models.Connection.receiver_id == payload.receiver_id)
        )
        | (
            (models.Connection.requester_id == payload.receiver_id)
            & (models.Connection.receiver_id == payload.requester_id)
        )
    ).first()
    if existing:
        raise HTTPException(status_code=400, detail="Connection already exists between these users")

    connection = models.Connection(**payload_dict(payload))
    db.add(connection)
    db.commit()
    db.refresh(connection)
    return {"connection_id": connection.connection_id, "status": connection.status}


@app.patch("/connections/{connection_id}", tags=["connections"])
def update_connection_status(
    connection_id: int,
    payload: ConnectionStatusUpdate,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    allowed_statuses = {"accepted", "rejected"}
    normalized_status = payload.status.lower()
    if normalized_status not in allowed_statuses:
        raise HTTPException(status_code=400, detail="Invalid connection status")

    connection = get_connection_by_id(db, connection_id)
    if current_user.user_id != connection.receiver_id and current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Only the receiving alumni can respond to this request")
    if connection.status != "pending":
        raise HTTPException(status_code=400, detail="Only pending requests can be updated")

    connection.status = normalized_status
    db.commit()
    # Award points to alumni for accepting connection
    if normalized_status == "accepted":
        _award_points(db, current_user.user_id, "connection", 5)
        db.commit()
    return {"connection_id": connection.connection_id, "status": connection.status}


@app.delete("/connections/{connection_id}", tags=["connections"])
def delete_connection(
    connection_id: int,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    connection = get_connection_by_id(db, connection_id)
    if current_user.user_id not in {connection.requester_id, connection.receiver_id} and current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Not authorized to delete this connection")
    db.delete(connection)
    db.commit()
    return {"status": "deleted"}


@app.post("/announcements")
def create_announcement(payload: AnnouncementCreate, db: Session = Depends(get_db)):
    user = get_user_by_id(db, payload.created_by)
    if user.role != "admin":
        raise HTTPException(status_code=400, detail="Only admin users can post announcements")
    announcement = models.Announcement(**payload_dict(payload))
    db.add(announcement)
    db.commit()
    db.refresh(announcement)
    return {"announcement_id": announcement.announcement_id, "message": "Announcement created"}


@app.get("/announcements")
def list_announcements(db: Session = Depends(get_db)):
    return db.query(models.Announcement).order_by(models.Announcement.created_at.desc()).all()


@app.patch("/announcements/{announcement_id}")
def update_announcement(announcement_id: int, payload: AnnouncementCreate, db: Session = Depends(get_db)):
    announcement = get_announcement_by_id(db, announcement_id)
    updates = payload_dict(payload)
    for field, value in updates.items():
        if value is not None:
            setattr(announcement, field, value)
    db.commit()
    return {"status": "updated"}


@app.delete("/announcements/{announcement_id}")
def delete_announcement(announcement_id: int, db: Session = Depends(get_db)):
    announcement = get_announcement_by_id(db, announcement_id)
    db.delete(announcement)
    db.commit()
    return {"status": "deleted"}


@app.get("/connections", tags=["connections"])
def list_connections(
    user_id: int,
    role: Optional[str] = None,
    status: Optional[str] = None,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    get_user_by_id(db, user_id)
    if current_user.user_id != user_id and current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Not authorized to view these connections")

    if role == "alumni":
        query = (
            db.query(models.Connection, models.User)
            .join(models.User, models.User.user_id == models.Connection.requester_id)
            .filter(models.Connection.receiver_id == user_id)
        )
        if status:
            query = query.filter(models.Connection.status == status.lower())
        else:
            query = query.filter(models.Connection.status == "pending")

        results = query.order_by(models.Connection.created_at.desc()).all()
        return [
            {
                "connection_id": conn.connection_id,
                "status": conn.status,
                "created_at": conn.created_at.isoformat(),
                "requester": {
                    "user_id": requester.user_id,
                    "full_name": requester.full_name,
                    "email": requester.email,
                    "role": requester.role,
                    "department": requester.department,
                    "graduation_year": requester.graduation_year,
                    "city": requester.city,
                    "profile_picture_url": requester.profile_picture_url,
                },
            }
            for conn, requester in results
        ]

    query = (
        db.query(models.Connection, models.User)
        .join(models.User, models.User.user_id == models.Connection.receiver_id)
        .filter(models.Connection.requester_id == user_id)
    )
    if status:
        query = query.filter(models.Connection.status == status.lower())

    results = query.order_by(models.Connection.created_at.desc()).all()
    return [
        {
            "connection_id": conn.connection_id,
            "status": conn.status,
            "created_at": conn.created_at.isoformat(),
            "receiver": {
                "user_id": receiver.user_id,
                "full_name": receiver.full_name,
                "email": receiver.email,
                "role": receiver.role,
                "department": receiver.department,
                "graduation_year": receiver.graduation_year,
                "city": receiver.city,
                "profile_picture_url": receiver.profile_picture_url,
            },
        }
        for conn, receiver in results
    ]


@app.post("/messages", tags=["messages"])
def create_message(payload: MessageCreate, db: Session = Depends(get_db)):
    sender = get_user_by_id(db, payload.sender_id)
    receiver = get_user_by_id(db, payload.receiver_id)
    ensure_connected(db, sender.user_id, receiver.user_id)

    message = models.Message(**payload_dict(payload))
    db.add(message)

    # Save to notification DB
    db.add(models.Notification(
        user_id=receiver.user_id,
        type="message",
        message=f"New message from {sender.full_name}: {payload.content[:80]}{'...' if len(payload.content) > 80 else ''}",
        is_read=False,
    ))

    db.commit()
    db.refresh(message)

    # Send FCM push in background thread
    if receiver.fcm_token:
        import threading
        threading.Thread(
            target=_send_fcm_push,
            args=(receiver.fcm_token, f"New message from {sender.full_name}", payload.content[:200]),
            daemon=True,
        ).start()

    return {"message_id": message.message_id, "sent_at": message.sent_at}


def _send_fcm_push(token: str, title: str, body: str, memo_id: str = ""):
    """Send FCM push notification via HTTP v1 API (DATA ONLY — no duplicates)."""
    try:
        import os
        import json
        import urllib.request
        from google.oauth2 import service_account
        import google.auth.transport.requests

        # Load service account
        sa_json_str = os.getenv("FIREBASE_SERVICE_ACCOUNT_JSON")
        sa_path = os.path.join(os.path.dirname(__file__), "firebase-service-account.json")

        if sa_json_str:
            try:
                sa_data = json.loads(sa_json_str.strip())
            except json.JSONDecodeError as e:
                print(f"[FCM] Invalid FIREBASE_SERVICE_ACCOUNT_JSON: {e}")
                print(f"[FCM] First 100 chars: {sa_json_str[:100]!r}")
                return
        elif os.path.exists(sa_path):
            with open(sa_path) as f:
                sa_data = json.load(f)
        else:
            print("[FCM] No service account configured")
            return

        credentials = service_account.Credentials.from_service_account_info(
            sa_data,
            scopes=["https://www.googleapis.com/auth/firebase.messaging"],
        )
        credentials.refresh(google.auth.transport.requests.Request())

        access_token = credentials.token
        project_id = sa_data["project_id"]

        url = f"https://fcm.googleapis.com/v1/projects/{project_id}/messages:send"

        # ✅ DATA ONLY PAYLOAD
        payload = json.dumps({
            "message": {
                "token": token,
                "data": {
                    "title": title,
                    "body": body[:200],
                    "memo_id": memo_id or "",
                },
                "android": {
                    "priority": "high"
                }
            }
        }).encode("utf-8")

        req = urllib.request.Request(
            url,
            data=payload,
            headers={
                "Authorization": f"Bearer {access_token}",
                "Content-Type": "application/json",
            },
        )

        with urllib.request.urlopen(req, timeout=10) as resp:
            result = resp.read().decode()
            print(f"[FCM] Sent: {result}")

    except Exception as e:
        print(f"[FCM] Error: {e}")


@app.get("/conversations", tags=["messages"])
def list_conversations(current_user: models.User = Depends(get_current_user), db: Session = Depends(get_db)):
    sender_ids = [m.receiver_id for m in db.query(models.Message).filter(models.Message.sender_id == current_user.user_id).all()]
    receiver_ids = [m.sender_id for m in db.query(models.Message).filter(models.Message.receiver_id == current_user.user_id).all()]

    unique_ids = list(set(sender_ids + receiver_ids))
    other_users = db.query(models.User).filter(models.User.user_id.in_(unique_ids)).all()

    result = []
    for u in other_users:
        user_data = serialize_user(u)
        # Count unread messages from this user to current user
        unread = db.query(models.Message).filter(
            models.Message.sender_id == u.user_id,
            models.Message.receiver_id == current_user.user_id,
            models.Message.is_read == False,
        ).count()
        user_data["unread_count"] = unread
        result.append(user_data)

    # Sort: conversations with unread messages first
    result.sort(key=lambda x: x["unread_count"], reverse=True)
    return result


@app.patch("/messages/{message_id}", tags=["messages"])
def edit_message(message_id: int, content: str = Body(..., embed=True), current_user: models.User = Depends(get_current_user), db: Session = Depends(get_db)):
    message = db.query(models.Message).filter(models.Message.message_id == message_id).first()
    if not message:
        raise HTTPException(status_code=404, detail="Message not found")
    
    if message.sender_id != current_user.user_id:
        raise HTTPException(status_code=403, detail="Not authorized to edit this message")

    now = datetime.now(timezone.utc)
    sent_at = message.sent_at.replace(tzinfo=timezone.utc) if message.sent_at.tzinfo is None else message.sent_at
    
    if now - sent_at > timedelta(minutes=5):
        raise HTTPException(status_code=400, detail="Edit period (5 mins) has expired")

    message.content = content
    db.commit()
    return {"status": "updated", "content": content}


@app.delete("/messages/clear", tags=["messages"])
def clear_chat(other_user_id: int, current_user: models.User = Depends(get_current_user), db: Session = Depends(get_db)):
    # Deletes all messages between current_user and another user
    messages = db.query(models.Message).filter(
        (
            (models.Message.sender_id == current_user.user_id)
            & (models.Message.receiver_id == other_user_id)
        )
        | (
            (models.Message.sender_id == other_user_id)
            & (models.Message.receiver_id == current_user.user_id)
        )
    ).all()
    
    count = 0
    for msg in messages:
        db.delete(msg)
        count += 1
        
    db.commit()
    return {"status": "cleared", "count": count}


@app.delete("/messages/{message_id}", tags=["messages"])
def delete_message(message_id: int, current_user: models.User = Depends(get_current_user), db: Session = Depends(get_db)):
    message = db.query(models.Message).filter(models.Message.message_id == message_id).first()
    if not message:
        raise HTTPException(status_code=404, detail="Message not found")
    
    if message.sender_id != current_user.user_id:
        raise HTTPException(status_code=403, detail="Not authorized to delete this message")

    now = datetime.now(timezone.utc)
    sent_at = message.sent_at.replace(tzinfo=timezone.utc) if message.sent_at.tzinfo is None else message.sent_at
    
    if now - sent_at > timedelta(hours=24):
        raise HTTPException(status_code=400, detail="Deletion period (24 hours) has expired")

    db.delete(message)
    db.commit()
    return {"status": "deleted"}


@app.post("/messages/bulk-delete", tags=["messages"])
def bulk_delete_messages(message_ids: list[int], current_user: models.User = Depends(get_current_user), db: Session = Depends(get_db)):
    now = datetime.now(timezone.utc)
    
    # Query only messages that belong to user and are < 24h old
    messages = db.query(models.Message).filter(
        models.Message.message_id.in_(message_ids),
        models.Message.sender_id == current_user.user_id
    ).all()
    
    deleted_count = 0
    for msg in messages:
        sent_at = msg.sent_at.replace(tzinfo=timezone.utc) if msg.sent_at.tzinfo is None else msg.sent_at
        if now - sent_at <= timedelta(hours=24):
            db.delete(msg)
            deleted_count += 1
            
    db.commit()
    return {"status": "deleted", "count": deleted_count}



@app.get("/messages", tags=["messages"])
def list_messages(user_id: int, other_user_id: int, current_user: models.User = Depends(get_current_user), db: Session = Depends(get_db)):
    get_user_by_id(db, user_id)
    get_user_by_id(db, other_user_id)
    
    if current_user.user_id != user_id and current_user.user_id != other_user_id and current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Not authorized to see these messages")

    ensure_connected(db, user_id, other_user_id)

    messages = db.query(models.Message).filter(
        (
            (models.Message.sender_id == user_id)
            & (models.Message.receiver_id == other_user_id)
        )
        | (
            (models.Message.sender_id == other_user_id)
            & (models.Message.receiver_id == user_id)
        )
    ).order_by(models.Message.sent_at.asc()).all()

    # Mark messages sent to current user as read
    db.query(models.Message).filter(
        models.Message.sender_id == other_user_id,
        models.Message.receiver_id == current_user.user_id,
        models.Message.is_read == False,
    ).update({"is_read": True})
    db.commit()

    return [
        {
            "message_id": m.message_id,
            "sender_id": m.sender_id,
            "receiver_id": m.receiver_id,
            "content": m.content,
            "sent_at": m.sent_at.isoformat(),
            "is_read": m.is_read,
        }
        for m in messages
    ]


@app.get("/messages/unread-count", tags=["messages"])
def get_unread_count(current_user: models.User = Depends(get_current_user), db: Session = Depends(get_db)):
    count = db.query(models.Message).filter(
        models.Message.receiver_id == current_user.user_id,
        models.Message.is_read == False,
    ).count()
    return {"unread_count": count}


@app.post("/resources", tags=["resources"])
def create_resource(payload: ResourceCreate, db: Session = Depends(get_db)):
    user = get_user_by_id(db, payload.created_by)
    if user.role not in {"staff", "admin"}:
        raise HTTPException(status_code=400, detail="Only staff or admin users can create resources")

    resource = models.Resource(**payload_dict(payload))
    db.add(resource)
    db.commit()
    db.refresh(resource)

    # Build notification content
    type_label = "📄 New Document" if payload.resource_type == "document" else "🔗 New Link"
    notif_title = f"{type_label}: {payload.title}"
    notif_body = payload.description or f"Shared by {user.full_name}"

    # Save to notification DB for all target users
    _notify_audience(
        db=db,
        target_role=payload.target_audience,
        sender_id=user.user_id,
        noti_type="broadcast",
        message=f"{notif_title} — {notif_body}",
    )

    # Send FCM push to target users who have a token
    import threading
    query = db.query(models.User).filter(models.User.fcm_token.isnot(None))
    if payload.target_audience != "all":
        query = query.filter(models.User.role == payload.target_audience)
    else:
        query = query.filter(models.User.role.in_(["student", "alumni"]))

    for target_user in query.all():
        if target_user.user_id == user.user_id:
            continue
        threading.Thread(
            target=_send_fcm_push,
            args=(target_user.fcm_token, notif_title, notif_body),
            daemon=True,
        ).start()

    return {"resource_id": resource.resource_id, "message": "Resource created"}


@app.get("/resources", tags=["resources"])
def list_resources(current_user: models.User = Depends(get_current_user), db: Session = Depends(get_db)):
    if current_user.role in {"admin", "staff"}:
        return db.query(models.Resource).order_by(models.Resource.created_at.desc()).all()
    
    return db.query(models.Resource).filter(
        (models.Resource.target_audience == "all") | (models.Resource.target_audience == current_user.role)
    ).order_by(models.Resource.created_at.desc()).all()


@app.delete("/resources/{resource_id}", tags=["resources"])
def delete_resource(resource_id: int, current_user: models.User = Depends(get_current_user), db: Session = Depends(get_db)):
    resource = db.query(models.Resource).filter(models.Resource.resource_id == resource_id).first()
    if not resource:
        raise HTTPException(status_code=404, detail="Resource not found")
    
    if current_user.role not in {"admin", "staff"} and resource.created_by != current_user.user_id:
        raise HTTPException(status_code=403, detail="Not authorized to delete this resource")

    db.delete(resource)
    db.commit()
    return {"status": "deleted"}


@app.post("/notifications")
def create_notification(payload: NotificationCreate, db: Session = Depends(get_db)):
    get_user_by_id(db, payload.user_id)
    notification = models.Notification(**payload_dict(payload))
    db.add(notification)
    db.commit()
    db.refresh(notification)
    return {"noti_id": notification.noti_id, "message": "Notification created"}


@app.get("/notifications/unread-count", tags=["events"])
def get_unread_notification_count(
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    count = db.query(models.Notification).filter(
        models.Notification.user_id == current_user.user_id,
        models.Notification.is_read == False,
    ).count()
    return {"unread_count": count}


@app.get("/notifications", tags=["events"])
def list_notifications(current_user: models.User = Depends(get_current_user), db: Session = Depends(get_db)):
    return db.query(models.Notification).filter(
        models.Notification.user_id == current_user.user_id
    ).order_by(models.Notification.created_at.desc()).all()


@app.patch("/notifications/mark-all-read", tags=["events"])
def mark_all_notifications_read(
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    db.query(models.Notification).filter(
        models.Notification.user_id == current_user.user_id,
        models.Notification.is_read == False,
    ).update({"is_read": True})
    db.commit()
    return {"status": "all_marked_read"}


@app.delete("/notifications/clear-all", tags=["events"])
def clear_all_notifications(
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    db.query(models.Notification).filter(
        models.Notification.user_id == current_user.user_id,
    ).delete(synchronize_session=False)
    db.commit()
    return {"status": "cleared"}


class BroadcastNotificationRequest(BaseModel):
    title: str
    message: str
    target_role: str  # "all", "student", "alumni", "staff"


@app.post("/notifications/broadcast", tags=["events"])
def broadcast_notification(
    payload: BroadcastNotificationRequest,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    if current_user.role not in {"admin", "staff"}:
        raise HTTPException(status_code=403, detail="Only admin or staff can broadcast notifications")

    query = db.query(models.User)
    if payload.target_role != "all":
        query = query.filter(models.User.role == payload.target_role)

    users = query.all()
    full_message = f"{payload.title}: {payload.message}" if payload.title else payload.message

    count = 0
    fcm_targets = []
    for user in users:
        if user.user_id == current_user.user_id:
            continue
        db.add(models.Notification(
            user_id=user.user_id,
            type="broadcast",
            message=full_message,
            is_read=False,
        ))
        if user.fcm_token:
            fcm_targets.append(user.fcm_token)
        count += 1

    db.add(models.BroadcastLog(
        sent_by=current_user.user_id,
        title=payload.title,
        message=payload.message,
        target_role=payload.target_role,
        recipient_count=count,
    ))
    db.commit()

    # Send FCM push to all target devices in background
    if fcm_targets:
        import threading
        for token in fcm_targets:
            threading.Thread(
                target=_send_fcm_push,
                args=(token, payload.title, payload.message),
                daemon=True,
            ).start()

    return {"status": "sent", "recipients": count}


@app.get("/notifications/sent", tags=["events"])
def get_sent_notifications(
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    if current_user.role not in {"admin", "staff"}:
        raise HTTPException(status_code=403, detail="Not authorized")

    logs = db.query(models.BroadcastLog).filter(
        models.BroadcastLog.sent_by == current_user.user_id
    ).order_by(models.BroadcastLog.created_at.desc()).all()

    return [
        {
            "log_id": log.log_id,
            "title": log.title,
            "message": log.message,
            "target_role": log.target_role,
            "recipient_count": log.recipient_count,
            "created_at": log.created_at.isoformat() if log.created_at else None,
        }
        for log in logs
    ]


class BroadcastLogDeleteRequest(BaseModel):
    log_ids: list[int]


@app.post("/notifications/sent/delete", tags=["events"])
def delete_broadcast_logs(
    payload: BroadcastLogDeleteRequest,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    if current_user.role not in {"admin", "staff"}:
        raise HTTPException(status_code=403, detail="Not authorized")

    db.query(models.BroadcastLog).filter(
        models.BroadcastLog.log_id.in_(payload.log_ids),
        models.BroadcastLog.sent_by == current_user.user_id,
    ).delete(synchronize_session=False)
    db.commit()
    return {"status": "deleted", "count": len(payload.log_ids)}


# Parameterized routes LAST — must come after all specific /notifications/* paths
@app.patch("/notifications/{noti_id}/read")
def mark_notification_as_read(noti_id: int, db: Session = Depends(get_db)):
    notification = get_notification_by_id(db, noti_id)
    notification.is_read = True
    db.commit()
    return {"status": "marked_as_read"}


@app.delete("/notifications/{noti_id}")
def delete_notification(noti_id: int, db: Session = Depends(get_db)):
    notification = get_notification_by_id(db, noti_id)
    db.delete(notification)
    db.commit()
    return {"status": "deleted"}

@app.on_event("startup")
def startup():
    init_db()


def custom_openapi():
    if app.openapi_schema:
        return app.openapi_schema
    schema = get_openapi(
        title=app.title,
        version=app.version,
        description=app.description,
        routes=app.routes,
        tags=tags_metadata,
    )
    schema.setdefault("components", {})
    schema["components"]["securitySchemes"] = {
        "BearerAuth": {
            "type": "http",
            "scheme": "bearer",
            "bearerFormat": "JWT",
            "description": "Enter your JWT token from POST /auth/login",
        }
    }
    for path in schema.get("paths", {}).values():
        for operation in path.values():
            if isinstance(operation, dict):
                operation.setdefault("security", [{"BearerAuth": []}])
    app.openapi_schema = schema
    return schema


app.openapi = custom_openapi


# ── Alumni Dashboard ────────────────────────────────────────────────────────

@app.get("/alumni/dashboard", tags=["alumni"])
def get_alumni_dashboard(
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    if current_user.role != "alumni":
        raise HTTPException(status_code=403, detail="Alumni only")

    uid = current_user.user_id

    # Internships posted
    internships_posted = db.query(models.Internship).filter(
        models.Internship.posted_by == uid
    ).count()

    # Total applicants across all my internships
    my_internship_ids = [i.internship_id for i in db.query(models.Internship.internship_id).filter(
        models.Internship.posted_by == uid
    ).all()]
    total_applicants = db.query(models.Application).filter(
        models.Application.internship_id.in_(my_internship_ids)
    ).count() if my_internship_ids else 0

    # New applicants (unread — Applied status)
    new_applicants = db.query(models.Application).filter(
        models.Application.internship_id.in_(my_internship_ids),
        models.Application.status == "Applied",
    ).count() if my_internship_ids else 0

    # Students connected
    from sqlalchemy import or_
    students_connected = db.query(models.Connection).filter(
        models.Connection.status == "accepted",
        or_(models.Connection.requester_id == uid, models.Connection.receiver_id == uid)
    ).count()

    # Pending connection requests
    pending_requests = db.query(models.Connection).filter(
        models.Connection.receiver_id == uid,
        models.Connection.status == "pending",
    ).count()

    # Unread messages
    unread_msgs = db.query(models.Message).filter(
        models.Message.receiver_id == uid,
        models.Message.is_read == False,
    ).count()

    # Points
    points_row = db.query(models.AlumniPoints).filter(
        models.AlumniPoints.alumni_id == uid
    ).first()
    points = points_row.points if points_row else 0

    # Recent notifications
    recent = db.query(models.Notification).filter(
        models.Notification.user_id == uid
    ).order_by(models.Notification.created_at.desc()).limit(5).all()

    # Pending connection requests detail
    pending_conns = db.query(models.Connection, models.User).join(
        models.User, models.User.user_id == models.Connection.requester_id
    ).filter(
        models.Connection.receiver_id == uid,
        models.Connection.status == "pending",
    ).order_by(models.Connection.created_at.desc()).limit(5).all()

    return {
        "stats": {
            "internships_posted": internships_posted,
            "students_connected": students_connected,
            "unread_messages": unread_msgs,
            "total_applicants": total_applicants,
            "new_applicants": new_applicants,
            "pending_requests": pending_requests,
            "points": points,
        },
        "pending_requests": [
            {
                "connection_id": c.connection_id,
                "student": {
                    "user_id": s.user_id,
                    "full_name": s.full_name,
                    "department": s.department,
                    "profile_picture_url": s.profile_picture_url,
                },
                "created_at": c.created_at.isoformat(),
            } for c, s in pending_conns
        ],
        "recent_activity": [
            {
                "message": r.message,
                "type": r.type,
                "created_at": r.created_at.isoformat() if r.created_at else None,
                "is_read": r.is_read,
            } for r in recent
        ],
    }


# ── Feedback ─────────────────────────────────────────────────────────────────

class FeedbackCreate(BaseModel):
    target_id: int
    target_role: str     # 'alumni' or 'staff'
    rating: int          # 1–5
    message: Optional[str] = None


def _serialize_feedback(f, include_target=True):
    d = {
        "feedback_id": f.feedback_id,
        "rating": f.rating,
        "message": f.message,
        "admin_response": f.admin_response,
        "target_role": f.target_role,
        "created_at": f.created_at.isoformat(),
        "student_name": f.student.full_name if f.student else "Unknown",
        "student_picture": f.student.profile_picture_url if f.student else None,
        "student_department": f.student.department if f.student else None,
    }
    if include_target:
        d["target_id"] = f.target_id
        d["target_name"] = f.target.full_name if f.target else "Unknown"
        d["target_picture"] = f.target.profile_picture_url if f.target else None
    return d


@app.post("/feedback", tags=["feedback"])
def submit_feedback(
    payload: FeedbackCreate,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    if current_user.role != "student":
        raise HTTPException(status_code=403, detail="Students only")
    if payload.target_role not in ("alumni", "staff"):
        raise HTTPException(status_code=400, detail="target_role must be 'alumni' or 'staff'")
    if not (1 <= payload.rating <= 5):
        raise HTTPException(status_code=400, detail="Rating must be 1–5")
    target = db.query(models.User).filter(models.User.user_id == payload.target_id).first()
    if not target or target.role != payload.target_role:
        raise HTTPException(status_code=404, detail="Target user not found")
    # One feedback per student per target — locked until deleted
    existing = db.query(models.Feedback).filter(
        models.Feedback.student_id == current_user.user_id,
        models.Feedback.target_id == payload.target_id,
    ).first()
    if existing:
        raise HTTPException(status_code=409, detail="You have already submitted feedback for this person. Delete it first to submit again.")
    fb = models.Feedback(
        student_id=current_user.user_id,
        target_id=payload.target_id,
        target_role=payload.target_role,
        rating=payload.rating,
        message=payload.message,
    )
    db.add(fb)
    db.commit()
    db.refresh(fb)
    return {"feedback_id": fb.feedback_id}


@app.delete("/feedback/{feedback_id}", tags=["feedback"])
def delete_feedback(
    feedback_id: int,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    fb = db.query(models.Feedback).filter(models.Feedback.feedback_id == feedback_id).first()
    if not fb:
        raise HTTPException(status_code=404, detail="Feedback not found")
    if fb.student_id != current_user.user_id:
        raise HTTPException(status_code=403, detail="Not your feedback")
    db.delete(fb)
    db.commit()
    return {"detail": "Deleted"}


@app.get("/feedback/sent", tags=["feedback"])
def get_sent_feedback(
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Student sees their own submitted feedbacks."""
    if current_user.role != "student":
        raise HTTPException(status_code=403, detail="Students only")
    rows = db.query(models.Feedback).filter(
        models.Feedback.student_id == current_user.user_id
    ).order_by(models.Feedback.created_at.desc()).all()
    return [_serialize_feedback(f) for f in rows]


@app.get("/feedback/received", tags=["feedback"])
def get_received_feedback(
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Alumni or staff sees feedback addressed to them."""
    if current_user.role not in ("alumni", "staff"):
        raise HTTPException(status_code=403, detail="Alumni or staff only")
    rows = db.query(models.Feedback).filter(
        models.Feedback.target_id == current_user.user_id
    ).order_by(models.Feedback.created_at.desc()).all()
    avg = round(sum(f.rating for f in rows) / len(rows), 1) if rows else 0
    return {
        "average_rating": avg,
        "total": len(rows),
        "feedbacks": [_serialize_feedback(f, include_target=False) for f in rows],
    }


@app.get("/feedback/all", tags=["feedback"])
def get_all_feedback(
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Admin sees all feedbacks."""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Admin only")
    rows = db.query(models.Feedback).order_by(models.Feedback.created_at.desc()).all()
    return [_serialize_feedback(f) for f in rows]


@app.patch("/feedback/{feedback_id}/respond", tags=["feedback"])
def respond_to_feedback(
    feedback_id: int,
    payload: dict,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Admin adds a response to a feedback."""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Admin only")
    fb = db.query(models.Feedback).filter(models.Feedback.feedback_id == feedback_id).first()
    if not fb:
        raise HTTPException(status_code=404, detail="Feedback not found")
    fb.admin_response = payload.get("response", "")
    db.commit()
    return {"detail": "Response saved"}


# ── Mentorship Slots ─────────────────────────────────────────────────────────

class MentorshipSlotCreate(BaseModel):
    day: str
    time_from: str
    time_to: str
    max_students: int = 3
    meeting_link: Optional[str] = None


@app.post("/mentorship/slots", tags=["alumni"])
def create_slot(
    payload: MentorshipSlotCreate,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    if current_user.role != "alumni":
        raise HTTPException(status_code=403, detail="Alumni only")
    slot = models.MentorshipSlot(
        alumni_id=current_user.user_id,
        day=payload.day,
        time_from=payload.time_from,
        time_to=payload.time_to,
        max_students=payload.max_students,
        meeting_link=payload.meeting_link,
    )
    db.add(slot)
    db.commit()
    db.refresh(slot)
    return {"slot_id": slot.slot_id}


@app.patch("/mentorship/slots/{slot_id}/meeting-link", tags=["alumni"])
def update_meeting_link(
    slot_id: int,
    payload: dict,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    if current_user.role != "alumni":
        raise HTTPException(status_code=403, detail="Alumni only")
    slot = db.query(models.MentorshipSlot).filter(models.MentorshipSlot.slot_id == slot_id).first()
    if not slot:
        raise HTTPException(status_code=404, detail="Slot not found")
    if slot.alumni_id != current_user.user_id:
        raise HTTPException(status_code=403, detail="Not your slot")
    slot.meeting_link = payload.get("meeting_link", "")
    db.commit()
    return {"detail": "Meeting link updated"}


@app.get("/mentorship/slots/{alumni_id}", tags=["alumni"])
def get_slots(alumni_id: int, db: Session = Depends(get_db)):
    slots = db.query(models.MentorshipSlot).filter(
        models.MentorshipSlot.alumni_id == alumni_id
    ).all()
    return [
        {
            "slot_id": s.slot_id,
            "day": s.day,
            "time_from": s.time_from,
            "time_to": s.time_to,
            "max_students": s.max_students,
            "meeting_link": s.meeting_link,
            "booked": len(s.bookings),
            "available": s.max_students - len(s.bookings),
        } for s in slots
    ]


@app.post("/mentorship/slots/{slot_id}/book", tags=["alumni"])
def book_slot(
    slot_id: int,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    if current_user.role != "student":
        raise HTTPException(status_code=403, detail="Students only")
    slot = db.query(models.MentorshipSlot).filter(models.MentorshipSlot.slot_id == slot_id).first()
    if not slot:
        raise HTTPException(status_code=404, detail="Slot not found")
    if len(slot.bookings) >= slot.max_students:
        raise HTTPException(status_code=400, detail="Slot is full")
    existing = db.query(models.MentorshipBooking).filter(
        models.MentorshipBooking.slot_id == slot_id,
        models.MentorshipBooking.student_id == current_user.user_id,
    ).first()
    if existing:
        raise HTTPException(status_code=400, detail="Already booked")
    booking = models.MentorshipBooking(slot_id=slot_id, student_id=current_user.user_id)
    db.add(booking)
    db.commit()
    return {"booking_id": booking.booking_id, "status": "pending"}


@app.delete("/mentorship/slots/{slot_id}", tags=["alumni"])
def delete_slot(
    slot_id: int,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    if current_user.role != "alumni":
        raise HTTPException(status_code=403, detail="Alumni only")
    slot = db.query(models.MentorshipSlot).filter(models.MentorshipSlot.slot_id == slot_id).first()
    if not slot:
        raise HTTPException(status_code=404, detail="Slot not found")
    if slot.alumni_id != current_user.user_id:
        raise HTTPException(status_code=403, detail="Not your slot")
    # Notify all booked students before deleting
    bookings = db.query(models.MentorshipBooking).filter(models.MentorshipBooking.slot_id == slot_id).all()
    for b in bookings:
        db.add(models.Notification(
            user_id=b.student_id,
            type="connection",
            message=f"The mentorship slot ({slot.day} {slot.time_from}–{slot.time_to}) by {current_user.full_name} has been cancelled.",
            is_read=False,
        ))
    db.query(models.MentorshipBooking).filter(models.MentorshipBooking.slot_id == slot_id).delete()
    db.delete(slot)
    db.commit()
    return {"detail": "Slot deleted"}


@app.get("/mentorship/slots/{slot_id}/students", tags=["alumni"])
def get_slot_students(
    slot_id: int,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    if current_user.role != "alumni":
        raise HTTPException(status_code=403, detail="Alumni only")
    slot = db.query(models.MentorshipSlot).filter(models.MentorshipSlot.slot_id == slot_id).first()
    if not slot:
        raise HTTPException(status_code=404, detail="Slot not found")
    if slot.alumni_id != current_user.user_id:
        raise HTTPException(status_code=403, detail="Not your slot")
    return [
        {
            "booking_id": b.booking_id,
            "status": b.status,
            "booked_at": b.created_at.isoformat() if b.created_at else None,
            "student_id": b.student.user_id,
            "student_name": b.student.full_name,
            "student_email": b.student.email,
            "student_department": b.student.department,
            "student_picture": b.student.profile_picture_url,
        }
        for b in slot.bookings
    ]


@app.patch("/mentorship/bookings/{booking_id}/status", tags=["alumni"])
def update_booking_status(
    booking_id: int,
    payload: dict,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    if current_user.role != "alumni":
        raise HTTPException(status_code=403, detail="Alumni only")
    booking = db.query(models.MentorshipBooking).filter(
        models.MentorshipBooking.booking_id == booking_id
    ).first()
    if not booking:
        raise HTTPException(status_code=404, detail="Booking not found")
    if booking.slot.alumni_id != current_user.user_id:
        raise HTTPException(status_code=403, detail="Not your slot")
    new_status = payload.get("status")
    if new_status not in ("confirmed", "rejected", "completed"):
        raise HTTPException(status_code=400, detail="Invalid status")

    # If accepting and a meeting_link is provided, save it to the slot
    if new_status == "confirmed":
        meeting_link = payload.get("meeting_link")
        if meeting_link:
            booking.slot.meeting_link = meeting_link

        # Notify the student
        db.add(models.Notification(
            user_id=booking.student_id,
            type="connection",
            message=f"Your mentorship session with {current_user.full_name} is confirmed! Join link is now available.",
            is_read=False,
        ))

    booking.status = new_status
    db.commit()
    return {"detail": "Status updated", "meeting_link": booking.slot.meeting_link}


@app.get("/mentorship/available", tags=["alumni"])
def get_available_slots(
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """All slots from all alumni with availability info, for students to browse."""
    slots = db.query(models.MentorshipSlot).all()
    result = []
    for s in slots:
        alumni_user = db.query(models.User).filter(models.User.user_id == s.alumni_id).first()
        alumni_profile = db.query(models.AlumniProfile).filter(
            models.AlumniProfile.alumni_id == s.alumni_id
        ).first()
        booked = len(s.bookings)
        available = s.max_students - booked
        # check if current student already booked this slot
        already_booked = any(b.student_id == current_user.user_id for b in s.bookings)
        result.append({
            "slot_id": s.slot_id,
            "day": s.day,
            "time_from": s.time_from,
            "time_to": s.time_to,
            "max_students": s.max_students,
            "booked": booked,
            "available": available,
            "already_booked": already_booked,
            "alumni_id": s.alumni_id,
            "alumni_name": alumni_user.full_name if alumni_user else "Unknown",
            "alumni_job_title": alumni_profile.job_title if alumni_profile else None,
            "alumni_company": alumni_profile.company if alumni_profile else None,
            "alumni_picture": alumni_user.profile_picture_url if alumni_user else None,
        })
    return result


@app.get("/mentorship/my-bookings", tags=["alumni"])
def get_my_bookings(
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Student's own bookings."""
    bookings = db.query(models.MentorshipBooking).filter(
        models.MentorshipBooking.student_id == current_user.user_id
    ).all()
    result = []
    for b in bookings:
        s = b.slot
        alumni_user = db.query(models.User).filter(models.User.user_id == s.alumni_id).first()
        alumni_profile = db.query(models.AlumniProfile).filter(
            models.AlumniProfile.alumni_id == s.alumni_id
        ).first()
        result.append({
            "booking_id": b.booking_id,
            "status": b.status,
            "slot_id": s.slot_id,
            "day": s.day,
            "time_from": s.time_from,
            "time_to": s.time_to,
            "alumni_id": s.alumni_id,
            "alumni_name": alumni_user.full_name if alumni_user else "Unknown",
            "alumni_job_title": alumni_profile.job_title if alumni_profile else None,
            "alumni_company": alumni_profile.company if alumni_profile else None,
            "alumni_picture": alumni_user.profile_picture_url if alumni_user else None,
            # Only expose meeting_link to confirmed students
            "meeting_link": s.meeting_link if b.status == "confirmed" else None,
            "has_feedback": db.query(models.Feedback).filter(
                models.Feedback.student_id == current_user.user_id,
                models.Feedback.target_id == s.alumni_id,
            ).first() is not None,
        })
    return result


# ── Points & Leaderboard ─────────────────────────────────────────────────────

def _award_points(db, alumni_id: int, action: str, pts: int):
    row = db.query(models.AlumniPoints).filter(models.AlumniPoints.alumni_id == alumni_id).first()
    if not row:
        row = models.AlumniPoints(alumni_id=alumni_id, points=0)
        db.add(row)
    row.points += pts
    if action == "internship":
        row.internships_posted += 1
    elif action == "answer":
        row.questions_answered += 1
    elif action == "mentorship":
        row.mentorships_completed += 1
    elif action == "connection":
        row.students_connected += 1


@app.get("/alumni/leaderboard", tags=["alumni"])
def get_leaderboard(db: Session = Depends(get_db)):
    rows = db.query(models.AlumniPoints, models.User).join(
        models.User, models.User.user_id == models.AlumniPoints.alumni_id
    ).order_by(models.AlumniPoints.points.desc()).limit(10).all()

    result = []
    for rank, (pts, user) in enumerate(rows, start=1):
        badges = _compute_badges(pts)
        result.append({
            "rank": rank,
            "user_id": user.user_id,
            "full_name": user.full_name,
            "profile_picture_url": user.profile_picture_url,
            "points": pts.points,
            "badges": badges,
            "internships_posted": pts.internships_posted,
            "questions_answered": pts.questions_answered,
            "mentorships_completed": pts.mentorships_completed,
        })
    return result


def _compute_badges(pts: models.AlumniPoints) -> list:
    badges = []
    if pts.mentorships_completed >= 20:
        badges.append({"name": "Top Mentor", "icon": "🏆"})
    elif pts.mentorships_completed >= 5:
        badges.append({"name": "Mentor", "icon": "🎓"})
    if pts.internships_posted >= 5:
        badges.append({"name": "Opportunity Giver", "icon": "💼"})
    if pts.questions_answered >= 10:
        badges.append({"name": "Active Contributor", "icon": "⭐"})
    if pts.students_connected >= 10:
        badges.append({"name": "Connector", "icon": "🤝"})
    if pts.points >= 100:
        badges.append({"name": "Alumni Star", "icon": "🌟"})
    return badges


@app.get("/alumni/my-points", tags=["alumni"])
def get_my_points(current_user: models.User = Depends(get_current_user), db: Session = Depends(get_db)):
    row = db.query(models.AlumniPoints).filter(models.AlumniPoints.alumni_id == current_user.user_id).first()
    if not row:
        return {"points": 0, "badges": [], "internships_posted": 0, "questions_answered": 0, "mentorships_completed": 0}
    return {
        "points": row.points,
        "badges": _compute_badges(row),
        "internships_posted": row.internships_posted,
        "questions_answered": row.questions_answered,
        "mentorships_completed": row.mentorships_completed,
        "students_connected": row.students_connected,
    }
